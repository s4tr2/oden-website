#!/usr/bin/env python3
"""
Generate the GTM Roadmap Template Excel file.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=15)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_widths(ws, widths):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def style_headers(ws, row, headers):
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_rows(ws, start_row, rows, height=34):
    for row_index, values in enumerate(rows, start=start_row):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = height


def create_quarterly_roadmap(wb):
    ws = wb.active
    ws.title = "Quarterly Roadmap"
    set_widths(ws, [26, 18, 16, 18, 10, 10, 10, 10, 14, 24, 28])
    ws["A1"] = "GTM ROADMAP - QUARTERLY VIEW"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Planning Period:"
    ws["B3"] = "[e.g. Q2 2026]"
    ws["D3"] = "Owner:"
    ws["E3"] = "[Name / Team]"

    style_headers(
        ws,
        5,
        [
            "Initiative",
            "Type",
            "Priority",
            "Owner",
            "Q1",
            "Q2",
            "Q3",
            "Q4",
            "Status",
            "Dependencies",
            "Notes",
        ],
    )

    rows = [
        ["Major Product Launch", "Launch", "High", "[PMM]", "", "X", "", "", "Planned", "Pricing, messaging sign-off", "Core cross-functional launch"],
        ["Enterprise Campaign", "Campaign", "High", "[Demand Gen]", "", "X", "X", "", "Planned", "Creative assets, audience list", "Supports pipeline target"],
        ["Sales Enablement Refresh", "Enablement", "Medium", "[Enablement]", "", "X", "", "", "In Progress", "Messaging updates", "Decks and training refresh"],
        ["Competitive Positioning Update", "Messaging", "Medium", "[PMM]", "X", "", "", "", "Planned", "Win/loss review", "Refine competitive narrative"],
        ["Segment Expansion", "Market Expansion", "High", "[GTM Lead]", "", "", "X", "X", "Planned", "ICP research, pricing review", "New vertical push"],
    ]
    style_rows(ws, 6, rows, height=40)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = "A5:K10"


def create_launch_calendar(wb):
    ws = wb.create_sheet("Launch Calendar")
    set_widths(ws, [26, 18, 18, 14, 24, 22, 28])
    ws["A1"] = "LAUNCH & CAMPAIGN CALENDAR"
    ws["A1"].font = TITLE_FONT
    style_headers(
        ws,
        3,
        [
            "Launch / Campaign",
            "Audience / Segment",
            "Channel / Motion",
            "Target Date",
            "Assets Needed",
            "Enablement Needed",
            "Readiness Notes",
        ],
    )
    rows = [
        ["AI Reporting Launch", "Existing customers", "PLG + lifecycle", "[2026-05-15]", "Landing page, email, blog", "Sales one-pager, FAQ", "Pending demo finalization"],
        ["Mid-Market Expansion Campaign", "Mid-market SaaS", "Paid + outbound", "[2026-06-01]", "Campaign page, ads, case study", "Persona talk track", "Depends on segment proof points"],
        ["Pricing & Packaging Update", "New prospects", "Sales-led", "[2026-06-20]", "Pricing page, enablement deck", "Objection handling update", "Leadership approval required"],
    ]
    style_rows(ws, 4, rows, height=40)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:G6"


def create_milestones_sheet(wb):
    ws = wb.create_sheet("Milestones")
    set_widths(ws, [26, 22, 14, 18, 18, 14, 24])
    ws["A1"] = "MILESTONES & DEPENDENCIES"
    ws["A1"].font = TITLE_FONT
    style_headers(
        ws,
        3,
        [
            "Milestone",
            "Related Initiative",
            "Due Date",
            "Owner",
            "Dependency Type",
            "Risk Level",
            "Next Step",
        ],
    )
    rows = [
        ["Messaging Approved", "AI Reporting Launch", "[2026-04-18]", "[PMM]", "Approval", "Medium", "Finalize review with leadership"],
        ["Pricing Finalized", "Pricing & Packaging Update", "[2026-05-02]", "[Finance / PMM]", "Cross-functional input", "High", "Resolve packaging options"],
        ["Sales Training Completed", "AI Reporting Launch", "[2026-05-10]", "[Enablement]", "Readiness", "Medium", "Schedule live session"],
        ["Campaign Creative Ready", "Mid-Market Expansion Campaign", "[2026-05-22]", "[Demand Gen]", "Asset dependency", "Low", "Approve final ad set"],
    ]
    style_rows(ws, 4, rows, height=38)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:G7"


def create_functional_owners_sheet(wb):
    ws = wb.create_sheet("Functional Owners")
    set_widths(ws, [24, 18, 18, 18, 18, 18, 18])
    ws["A1"] = "FUNCTIONAL OWNERS"
    ws["A1"].font = TITLE_FONT
    style_headers(
        ws,
        3,
        [
            "Initiative",
            "Product Marketing",
            "Marketing",
            "Sales / Enablement",
            "Product",
            "CS / Support",
            "Executive Sponsor",
        ],
    )
    rows = [
        ["AI Reporting Launch", "[PMM]", "[Demand Gen]", "[Enablement]", "[PM]", "[CS Lead]", "[VP Marketing]"],
        ["Mid-Market Expansion Campaign", "[PMM]", "[Demand Gen]", "[Sales Lead]", "[PM]", "[CS Lead]", "[GTM Leader]"],
        ["Pricing & Packaging Update", "[PMM]", "[Lifecycle]", "[Sales Ops]", "[Product]", "[Support]", "[CRO]"],
    ]
    style_rows(ws, 4, rows, height=32)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:G6"


def create_status_tracker(wb):
    ws = wb.create_sheet("Status Tracker")
    set_widths(ws, [24, 14, 14, 20, 24, 26])
    ws["A1"] = "STATUS TRACKER"
    ws["A1"].font = TITLE_FONT
    style_headers(
        ws,
        3,
        [
            "Initiative",
            "Status",
            "Confidence",
            "Upcoming Milestone",
            "Key Risk",
            "Action Needed",
        ],
    )
    rows = [
        ["AI Reporting Launch", "In Progress", "Medium", "Sales training", "Demo not finalized", "Approve final demo flow"],
        ["Mid-Market Expansion Campaign", "Planned", "High", "Creative ready", "Case study still pending", "Lock customer proof point"],
        ["Pricing & Packaging Update", "At Risk", "Low", "Pricing finalized", "Leadership alignment", "Escalate packaging decision"],
    ]
    style_rows(ws, 4, rows, height=36)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:F6"


def main():
    wb = Workbook()
    create_quarterly_roadmap(wb)
    create_launch_calendar(wb)
    create_milestones_sheet(wb)
    create_functional_owners_sheet(wb)
    create_status_tracker(wb)

    output_path = os.path.join(
        os.path.dirname(__file__), "..", "public", "templates", "gtm-roadmap-template.xlsx"
    )
    wb.save(output_path)
    print(f"Created {output_path}")


if __name__ == "__main__":
    main()
