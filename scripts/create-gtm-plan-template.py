#!/usr/bin/env python3
"""
Script to generate the GTM Plan Template Excel file.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)

# Border style
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def style_header_row(ws, row, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER


def style_section_header(ws, row, col, text):
    """Create a section header cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = SECTION_FILL
    cell.font = BOLD_FONT
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER
    return cell


def create_strategic_overview_sheet(wb):
    """Create the Strategic Overview sheet."""
    ws = wb.active
    ws.title = "Strategic Overview"

    set_column_widths(ws, [25, 60])

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="GTM PLAN - STRATEGIC OVERVIEW")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    # Plan Info
    ws.cell(row=row, column=1, value="Plan Period:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="[Year/Quarter, e.g., 2026 Annual Plan]")
    row += 1
    ws.cell(row=row, column=1, value="Last Updated:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="[Date]")
    row += 1
    ws.cell(row=row, column=1, value="Owner:").font = BOLD_FONT
    ws.cell(row=row, column=2, value="[Name/Team]")
    row += 2

    # Section 1: Business Objectives
    ws.merge_cells(f"A{row}:B{row}")
    style_section_header(ws, row, 1, "BUSINESS OBJECTIVES")
    row += 1

    objectives = [
        "Revenue Target: [e.g., $XM ARR by end of year]",
        "Customer Acquisition: [e.g., X new customers]",
        "Market Expansion: [e.g., enter X new segments/geos]",
        "Product Goals: [e.g., launch X major features]",
        "Brand Awareness: [e.g., achieve X% unaided awareness]",
    ]

    for obj in objectives:
        ws.cell(row=row, column=1, value=f"• {obj}").border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B{row}:B{row}")
        ws.row_dimensions[row].height = 25
        row += 1

    row += 1

    # Section 2: Target Market
    ws.merge_cells(f"A{row}:B{row}")
    style_section_header(ws, row, 1, "TARGET MARKET")
    row += 1

    market_fields = [
        ("Primary Segment", "[e.g., Mid-market SaaS companies, 100-500 employees]"),
        ("Secondary Segment", "[e.g., Enterprise, 500+ employees]"),
        ("Geographic Focus", "[e.g., North America, EMEA, APAC]"),
        ("Ideal Customer Profile", "[e.g., B2B SaaS companies with $10M+ revenue]"),
        ("Buyer Personas", "[e.g., VP of Product, CMO, Head of Growth]"),
    ]

    for field, value in market_fields:
        ws.cell(row=row, column=1, value=field).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Section 3: Positioning & Messaging
    ws.merge_cells(f"A{row}:B{row}")
    style_section_header(ws, row, 1, "POSITIONING & MESSAGING")
    row += 1

    positioning_fields = [
        ("Value Proposition", "[Core value prop statement]"),
        ("Key Messages", "[3-4 key messages by persona]"),
        ("Competitive Differentiation", "[How we're different from competitors]"),
        ("Proof Points", "[Customer evidence, metrics, case studies]"),
    ]

    for field, value in positioning_fields:
        ws.cell(row=row, column=1, value=field).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 40
        row += 1

    row += 1

    # Section 4: Go-To-Market Strategy
    ws.merge_cells(f"A{row}:B{row}")
    style_section_header(ws, row, 1, "GO-TO-MARKET STRATEGY")
    row += 1

    strategy_fields = [
        ("Primary Channels", "[e.g., Product-led growth, Inside sales, Partnerships]"),
        ("Sales Motion", "[e.g., Self-serve, Sales-assisted, Enterprise sales]"),
        ("Pricing Strategy", "[e.g., Freemium, Usage-based, Seat-based]"),
        ("Partnership Strategy", "[e.g., Channel partners, Integrations, Alliances]"),
    ]

    for field, value in strategy_fields:
        ws.cell(row=row, column=1, value=field).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1


def create_roadmap_timeline_sheet(wb):
    """Create the Strategic Initiative Timeline sheet."""
    ws = wb.create_sheet("Strategic Initiative Timeline")

    set_column_widths(ws, [30, 50, 15, 15, 15, 15, 20, 20])

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="STRATEGIC INITIATIVE TIMELINE")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    headers = [
        "Initiative/Activity",
        "Description",
        "Q1",
        "Q2",
        "Q3",
        "Q4",
        "Owner",
        "Status",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    # Q1 Section
    ws.merge_cells(f"A{row}:H{row}")
    style_section_header(ws, row, 1, "Q1 INITIATIVES")
    row += 1

    q1_initiatives = [
        (
            "Product Launch: [Feature Name]",
            "Launch major feature with full GTM support",
            "X",
            "",
            "",
            "",
            "[Product Marketing]",
            "Planned",
        ),
        (
            "Market Expansion: [Segment/Geo]",
            "Enter new market segment or geography",
            "X",
            "",
            "",
            "",
            "[GTM Lead]",
            "Planned",
        ),
        (
            "Brand Campaign: [Campaign Name]",
            "Launch brand awareness campaign",
            "X",
            "",
            "",
            "",
            "[Marketing]",
            "Planned",
        ),
        (
            "Sales Enablement: [Program]",
            "New sales training/enablement program",
            "X",
            "",
            "",
            "",
            "[Sales Enablement]",
            "Planned",
        ),
    ]

    for initiative, desc, q1, q2, q3, q4, owner, status in q1_initiatives:
        ws.cell(row=row, column=1, value=initiative).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=q3).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=q4).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=8, value=status).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Q2 Section
    ws.merge_cells(f"A{row}:H{row}")
    style_section_header(ws, row, 1, "Q2 INITIATIVES")
    row += 1

    q2_initiatives = [
        (
            "Product Launch: [Feature Name]",
            "Launch next major feature",
            "",
            "X",
            "",
            "",
            "[Product Marketing]",
            "Planned",
        ),
        (
            "Partnership Launch: [Partner]",
            "Launch strategic partnership",
            "",
            "X",
            "",
            "",
            "[Partnerships]",
            "Planned",
        ),
        (
            "Content Campaign: [Topic]",
            "Major content marketing campaign",
            "",
            "X",
            "",
            "",
            "[Content Marketing]",
            "Planned",
        ),
    ]

    for initiative, desc, q1, q2, q3, q4, owner, status in q2_initiatives:
        ws.cell(row=row, column=1, value=initiative).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=q3).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=q4).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=8, value=status).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Q3 Section
    ws.merge_cells(f"A{row}:H{row}")
    style_section_header(ws, row, 1, "Q3 INITIATIVES")
    row += 1

    q3_initiatives = [
        (
            "Product Launch: [Feature Name]",
            "Launch major feature",
            "",
            "",
            "X",
            "",
            "[Product Marketing]",
            "Planned",
        ),
        (
            "Event: [Conference/Webinar]",
            "Major industry event or webinar series",
            "",
            "",
            "X",
            "",
            "[Events Marketing]",
            "Planned",
        ),
    ]

    for initiative, desc, q1, q2, q3, q4, owner, status in q3_initiatives:
        ws.cell(row=row, column=1, value=initiative).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=q3).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=q4).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=8, value=status).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Q4 Section
    ws.merge_cells(f"A{row}:H{row}")
    style_section_header(ws, row, 1, "Q4 INITIATIVES")
    row += 1

    q4_initiatives = [
        (
            "Product Launch: [Feature Name]",
            "Launch major feature",
            "",
            "",
            "",
            "X",
            "[Product Marketing]",
            "Planned",
        ),
        (
            "Annual Planning",
            "2027 GTM planning and strategy",
            "",
            "",
            "",
            "X",
            "[GTM Lead]",
            "Planned",
        ),
    ]

    for initiative, desc, q1, q2, q3, q4, owner, status in q4_initiatives:
        ws.cell(row=row, column=1, value=initiative).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=q3).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=q4).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=8, value=status).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1


def create_channel_strategy_sheet(wb):
    """Create the Channel Strategy sheet."""
    ws = wb.create_sheet("Channel Strategy")

    set_column_widths(ws, [25, 30, 20, 20, 20, 30])

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="CHANNEL STRATEGY & BUDGET")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    headers = [
        "Channel",
        "Strategy/Tactics",
        "Budget Allocation",
        "Q1 Budget",
        "Q2-Q4 Budget",
        "Success Metrics",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    channels = [
        (
            "Digital Marketing",
            "SEO, SEM, Content Marketing, Social Ads",
            "X%",
            "$X",
            "$X",
            "Website traffic, MQLs, CAC",
        ),
        (
            "Product-Led Growth",
            "In-product onboarding, feature adoption",
            "X%",
            "$X",
            "$X",
            "Sign-ups, activation rate, conversion",
        ),
        (
            "Sales & SDR",
            "Outbound, inbound sales, SDR team",
            "X%",
            "$X",
            "$X",
            "SQLs, pipeline, win rate",
        ),
        (
            "Partnerships",
            "Channel partners, integrations, alliances",
            "X%",
            "$X",
            "$X",
            "Partner-sourced revenue, deals",
        ),
        (
            "Events & Webinars",
            "Conferences, trade shows, virtual events",
            "X%",
            "$X",
            "$X",
            "Event leads, pipeline generated",
        ),
        (
            "Content Marketing",
            "Blog, eBooks, webinars, case studies",
            "X%",
            "$X",
            "$X",
            "Content engagement, downloads, MQLs",
        ),
        (
            "PR & Analyst Relations",
            "Media relations, analyst briefings",
            "X%",
            "$X",
            "$X",
            "Press mentions, analyst coverage",
        ),
        (
            "Customer Marketing",
            "Referrals, case studies, advocacy",
            "X%",
            "$X",
            "$X",
            "NPS, referral rate, case studies",
        ),
    ]

    for channel, strategy, alloc, q1, q2_4, metrics in channels:
        ws.cell(row=row, column=1, value=channel).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=strategy).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=alloc).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=5, value=q2_4).border = THIN_BORDER
        ws.cell(row=row, column=6, value=metrics).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 35
        row += 1

    row += 1

    # Budget Summary
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "BUDGET SUMMARY")
    row += 1

    summary_fields = [
        ("Total GTM Budget", "$X"),
        ("Q1 Budget", "$X"),
        ("Q2-Q4 Budget", "$X"),
        ("Budget per Channel (Avg)", "$X"),
    ]

    for field, value in summary_fields:
        ws.cell(row=row, column=1, value=field).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).border = THIN_BORDER
        ws.cell(row=row, column=2).font = BOLD_FONT
        row += 1


def create_kpis_metrics_sheet(wb):
    """Create the KPIs & Metrics sheet."""
    ws = wb.create_sheet("KPIs & Metrics")

    set_column_widths(ws, [30, 20, 20, 20, 20, 30])

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="KPIs & SUCCESS METRICS")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    headers = [
        "Metric",
        "Baseline",
        "Q1 Target",
        "Q2 Target",
        "Q3-Q4 Target",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    # Revenue Metrics
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "REVENUE METRICS")
    row += 1

    revenue_metrics = [
        ("ARR/MRR", "", "", "", "", "Annual/Monthly Recurring Revenue"),
        ("New Customers", "", "", "", "", "Number of new customers acquired"),
        ("Expansion Revenue", "", "", "", "", "Upsell/cross-sell revenue"),
        ("Average Deal Size", "", "", "", "", "ACV or deal value"),
    ]

    for metric, baseline, q1, q2, q3_4, notes in revenue_metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=baseline).border = THIN_BORDER
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=5, value=q3_4).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1

    row += 1

    # Marketing Metrics
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "MARKETING METRICS")
    row += 1

    marketing_metrics = [
        ("MQLs (Marketing Qualified Leads)", "", "", "", "", "Leads from marketing"),
        ("SQLs (Sales Qualified Leads)", "", "", "", "", "Leads accepted by sales"),
        ("Website Traffic", "", "", "", "", "Unique visitors"),
        ("Content Engagement", "", "", "", "", "Downloads, views, shares"),
        ("Brand Awareness", "", "", "", "", "Survey-based awareness"),
    ]

    for metric, baseline, q1, q2, q3_4, notes in marketing_metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=baseline).border = THIN_BORDER
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=5, value=q3_4).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1

    row += 1

    # Sales Metrics
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "SALES METRICS")
    row += 1

    sales_metrics = [
        ("Pipeline Value", "", "", "", "", "Total pipeline $ value"),
        ("Win Rate", "", "", "", "", "Deals won / total deals"),
        ("Sales Cycle Length", "", "", "", "", "Days from SQL to close"),
        ("CAC (Customer Acquisition Cost)", "", "", "", "", "Total acquisition cost / customers"),
        ("LTV:CAC Ratio", "", "", "", "", "Lifetime value to CAC ratio"),
    ]

    for metric, baseline, q1, q2, q3_4, notes in sales_metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=baseline).border = THIN_BORDER
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=5, value=q3_4).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1

    row += 1

    # Product Metrics
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "PRODUCT METRICS")
    row += 1

    product_metrics = [
        ("Product Adoption Rate", "", "", "", "", "% of customers using feature"),
        ("Feature Usage", "", "", "", "", "DAU/MAU, feature engagement"),
        ("Time to Value", "", "", "", "", "Days to first value realization"),
        ("Product-Qualified Leads", "", "", "", "", "PQLs from product usage"),
    ]

    for metric, baseline, q1, q2, q3_4, notes in product_metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=baseline).border = THIN_BORDER
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=5, value=q3_4).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1


def create_resource_planning_sheet(wb):
    """Create the Resource Planning sheet."""
    ws = wb.create_sheet("Resource Planning")

    set_column_widths(ws, [25, 20, 20, 20, 20, 30])

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="RESOURCE PLANNING & TEAM")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    headers = [
        "Team/Role",
        "Current Headcount",
        "Q1 Plan",
        "Q2 Plan",
        "Q3-Q4 Plan",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    teams = [
        (
            "Product Marketing",
            "X",
            "X",
            "X",
            "X",
            "Owns GTM strategy and execution",
        ),
        (
            "Marketing",
            "X",
            "X",
            "X",
            "X",
            "Demand gen, content, brand",
        ),
        (
            "Sales",
            "X",
            "X",
            "X",
            "X",
            "Account executives, SDRs",
        ),
        (
            "Sales Enablement",
            "X",
            "X",
            "X",
            "X",
            "Training and enablement",
        ),
        (
            "Customer Success",
            "X",
            "X",
            "X",
            "X",
            "Onboarding and retention",
        ),
        (
            "Product",
            "X",
            "X",
            "X",
            "X",
            "Product managers, designers",
        ),
        (
            "Partnerships",
            "X",
            "X",
            "X",
            "X",
            "Channel and partner management",
        ),
    ]

    for team, current, q1, q2, q3_4, notes in teams:
        ws.cell(row=row, column=1, value=team).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=current).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=q1).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=q2).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=q3_4).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1


def create_risks_assumptions_sheet(wb):
    """Create the Risks & Assumptions sheet."""
    ws = wb.create_sheet("Risks & Assumptions")

    set_column_widths(ws, [25, 40, 30, 30])

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="RISKS, ASSUMPTIONS & MITIGATION")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    headers = [
        "Risk/Assumption",
        "Description",
        "Impact",
        "Mitigation Plan",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    risks = [
        (
            "Market Conditions",
            "Economic downturn affects buyer budgets",
            "High",
            "Focus on ROI messaging, adjust pricing if needed",
        ),
        (
            "Competitive Response",
            "Competitor launches similar feature",
            "Medium",
            "Accelerate differentiation messaging, highlight unique value",
        ),
        (
            "Product Delays",
            "Key features delayed, affecting launch timeline",
            "High",
            "Build buffer time, have backup launch content ready",
        ),
        (
            "Resource Constraints",
            "Team capacity limits execution",
            "Medium",
            "Prioritize initiatives, consider contractors/agencies",
        ),
        (
            "Channel Performance",
            "Primary channel underperforms",
            "Medium",
            "Diversify channels, have backup plans ready",
        ),
        (
            "Messaging Assumption",
            "Target audience responds differently than expected",
            "Low",
            "Test messaging early, iterate based on feedback",
        ),
    ]

    for risk, desc, impact, mitigation in risks:
        ws.cell(row=row, column=1, value=risk).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=impact).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=mitigation).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 35
        row += 1


def main():
    """Generate the GTM plan/roadmap template Excel file."""
    wb = Workbook()

    create_strategic_overview_sheet(wb)
    create_roadmap_timeline_sheet(wb)
    create_channel_strategy_sheet(wb)
    create_kpis_metrics_sheet(wb)
    create_resource_planning_sheet(wb)
    create_risks_assumptions_sheet(wb)

    # Save to public/templates
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(
        script_dir, "..", "public", "templates", "gtm-plan-template.xlsx"
    )
    output_path = os.path.normpath(output_path)

    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
