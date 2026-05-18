#!/usr/bin/env python3
"""
Script to generate the Battle Card Template Excel file.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)

# Border style
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def style_header_row(ws, row, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER


def style_section_header(ws, row, col, text):
    """Create a section header cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = SECTION_FILL
    cell.font = BOLD_FONT
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER
    return cell


def create_battle_card_sheet(wb):
    """Create the main Battle Card Template sheet."""
    ws = wb.active
    ws.title = "Battle Card Template"

    set_column_widths(ws, [25, 50, 25, 50])

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="BATTLE CARD: [COMPETITOR NAME]")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    # Last Updated
    ws.cell(row=2, column=1, value="Last Updated:")
    ws.cell(row=2, column=2, value="[Date]")
    ws.cell(row=2, column=3, value="Owner:")
    ws.cell(row=2, column=4, value="[Name]")

    row = 4

    # Section 1: Competitor Snapshot
    ws.merge_cells(f"A{row}:D{row}")
    style_section_header(ws, row, 1, "COMPETITOR SNAPSHOT")
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    snapshot_fields = [
        ("Company Name", "", "Founded", ""),
        ("Headquarters", "", "Employees", ""),
        ("Funding/Revenue", "", "Website", ""),
        ("Target Market", "", "Primary Use Case", ""),
        ("Positioning Statement", "", "", ""),
    ]

    for field_row in snapshot_fields:
        ws.cell(row=row, column=1, value=field_row[0]).font = BOLD_FONT
        ws.cell(row=row, column=2, value=field_row[1])
        ws.cell(row=row, column=3, value=field_row[2]).font = BOLD_FONT
        ws.cell(row=row, column=4, value=field_row[3])
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1

    # Section 2: Feature Comparison
    ws.merge_cells(f"A{row}:D{row}")
    style_section_header(ws, row, 1, "FEATURE COMPARISON")
    row += 1

    headers = ["Feature", "Us", "Them", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = LIGHT_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    row += 1

    features = [
        "[Feature 1]",
        "[Feature 2]",
        "[Feature 3]",
        "[Feature 4]",
        "[Feature 5]",
        "[Feature 6]",
    ]
    for feature in features:
        ws.cell(row=row, column=1, value=feature).border = THIN_BORDER
        ws.cell(row=row, column=2, value="✓ / ✗ / Partial").border = THIN_BORDER
        ws.cell(row=row, column=3, value="✓ / ✗ / Partial").border = THIN_BORDER
        ws.cell(row=row, column=4, value="").border = THIN_BORDER
        row += 1

    row += 1

    # Section 3: Their Strengths & Weaknesses
    ws.merge_cells(f"A{row}:B{row}")
    style_section_header(ws, row, 1, "THEIR STRENGTHS")
    ws.merge_cells(f"C{row}:D{row}")
    style_section_header(ws, row, 3, "THEIR WEAKNESSES")
    row += 1

    for i in range(4):
        ws.merge_cells(f"A{row}:B{row}")
        ws.cell(row=row, column=1, value=f"• [Strength {i+1}]").border = THIN_BORDER
        ws.merge_cells(f"C{row}:D{row}")
        ws.cell(row=row, column=3, value=f"• [Weakness {i+1}]").border = THIN_BORDER
        row += 1

    row += 1

    # Section 4: Common Objections & Responses
    ws.merge_cells(f"A{row}:D{row}")
    style_section_header(ws, row, 1, "COMMON OBJECTIONS & RESPONSES")
    row += 1

    headers = ["Objection", "Response"]
    ws.cell(row=row, column=1, value="Objection").fill = LIGHT_FILL
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.merge_cells(f"B{row}:D{row}")
    ws.cell(row=row, column=2, value="Response").fill = LIGHT_FILL
    ws.cell(row=row, column=2).font = BOLD_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1

    objections = [
        (
            '"They have [feature] that you don\'t"',
            "[Acknowledge, reframe, provide alternative or roadmap]",
        ),
        (
            '"They\'re cheaper"',
            "[Focus on value, TCO, hidden costs, or what's included]",
        ),
        (
            '"They integrate with [tool] better"',
            "[Highlight your integrations, API flexibility, or partnership]",
        ),
        (
            '"They\'re more established"',
            "[Focus on innovation, customer success, or specific advantages]",
        ),
    ]

    for obj, resp in objections:
        ws.cell(row=row, column=1, value=obj).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=resp).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 40
        row += 1

    row += 1

    # Section 5: Win Themes
    ws.merge_cells(f"A{row}:D{row}")
    style_section_header(ws, row, 1, "WIN THEMES (Key differentiators to emphasize)")
    row += 1

    for i in range(3):
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(
            row=row,
            column=1,
            value=f"{i+1}. [Win theme that resonates with buyers vs this competitor]",
        )
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 25
        row += 1

    row += 1

    # Section 6: Landmines
    ws.merge_cells(f"A{row}:D{row}")
    style_section_header(ws, row, 1, "LANDMINES (Questions to plant with the buyer)")
    row += 1

    landmines = [
        '"Ask them about their uptime SLA and what happens when they miss it"',
        '"Request a reference customer in your industry with similar scale"',
        '"Ask how long their typical implementation takes end-to-end"',
        '"Ask about their pricing for [feature] - it may be an add-on"',
    ]

    for landmine in landmines:
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1, value=f"• {landmine}")
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # Section 7: Proof Points
    ws.merge_cells(f"A{row}:D{row}")
    style_section_header(ws, row, 1, "PROOF POINTS & CUSTOMER EVIDENCE")
    row += 1

    headers = ["Type", "Details"]
    ws.cell(row=row, column=1, value="Type").fill = LIGHT_FILL
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.merge_cells(f"B{row}:D{row}")
    ws.cell(row=row, column=2, value="Details").fill = LIGHT_FILL
    ws.cell(row=row, column=2).font = BOLD_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1

    proof_points = [
        ("Customer Quote", '"[Quote from customer who switched from competitor]"'),
        ("Case Study", "[Customer name] achieved [X% improvement] in [metric]"),
        ("Analyst Report", "[Analyst firm] rated us [rating] in [category]"),
        ("Review Site", "[X] rating on [G2/Capterra] vs their [Y] rating"),
    ]

    for ptype, details in proof_points:
        ws.cell(row=row, column=1, value=ptype).border = THIN_BORDER
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=details).border = THIN_BORDER
        row += 1


def create_competitor_tracker_sheet(wb):
    """Create the Competitor Tracker sheet."""
    ws = wb.create_sheet("Competitor Tracker")

    set_column_widths(ws, [25, 15, 15, 30, 35, 15, 15, 20])

    headers = [
        "Competitor",
        "Category",
        "Threat Level",
        "Market Positioning",
        "Key Differentiators",
        "Last Updated",
        "Battle Card Status",
        "Owner",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Sample data rows
    sample_data = [
        (
            "[Competitor A]",
            "Direct",
            "High",
            "Enterprise-focused, premium pricing",
            "Strong brand, large customer base",
            "",
            "Draft",
            "",
        ),
        (
            "[Competitor B]",
            "Direct",
            "Medium",
            "SMB-focused, low-cost leader",
            "Easy to use, quick setup",
            "",
            "Complete",
            "",
        ),
        (
            "[Competitor C]",
            "Indirect",
            "Low",
            "Adjacent market, different use case",
            "Specialized features for [niche]",
            "",
            "Not Started",
            "",
        ),
        (
            "[Competitor D]",
            "Direct",
            "High",
            "Mid-market, balanced offering",
            "Good integrations, strong support",
            "",
            "In Progress",
            "",
        ),
    ]

    for row_num, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 35

    # Add dropdown-style guidance in a note
    ws.cell(row=7, column=1, value="Threat Level options:").font = BOLD_FONT
    ws.cell(row=7, column=2, value="High / Medium / Low")
    ws.cell(row=8, column=1, value="Status options:").font = BOLD_FONT
    ws.cell(
        row=8,
        column=2,
        value="Not Started / Draft / In Progress / Complete / Needs Update",
    )


def create_objection_library_sheet(wb):
    """Create the Objection Library sheet."""
    ws = wb.create_sheet("Objection Library")

    set_column_widths(ws, [15, 35, 50, 35, 20])

    headers = [
        "Category",
        "Objection",
        "Recommended Response",
        "Supporting Evidence",
        "Competitors",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Sample objections
    objections = [
        (
            "Price",
            '"You\'re more expensive than [Competitor]"',
            'Focus on total value: "While our list price may be higher, customers typically see [X%] lower TCO because [we include support/training/features they charge extra for]. Let me show you a breakdown..."',
            "TCO calculator, customer ROI case study",
            "All",
        ),
        (
            "Features",
            '"You don\'t have [specific feature]"',
            'Acknowledge and reframe: "You\'re right that we approach this differently. Instead of [their approach], we [your approach] because [reason]. Customers tell us this works better because..."',
            "Customer testimonial, product roadmap (if coming soon)",
            "[Competitor A]",
        ),
        (
            "Company",
            '"You\'re smaller/newer than [Competitor]"',
            "\"That's true, and here's why that's an advantage: [faster innovation / more responsive support / you're not just a number]. We have [X] customers including [notable names]...\"",
            "Customer logos, growth metrics, support response time",
            "[Competitor B], [Competitor C]",
        ),
        (
            "Integration",
            '"You don\'t integrate with [tool]"',
            '"We integrate with [X tools] natively and have a robust API. For [specific tool], customers typically use [workaround/Zapier/etc]. Is that integration critical for your decision?"',
            "Integration directory link, API documentation",
            "[Competitor A]",
        ),
        (
            "Risk",
            '"What if your company doesn\'t make it?"',
            '"Fair concern. We have [X years runway / are profitable / backed by Y investors]. We also offer [data export / open standards / escrow] so you\'re never locked in."',
            "Funding announcement, profitability metrics, data portability docs",
            "All (especially vs large incumbents)",
        ),
    ]

    for row_num, data in enumerate(objections, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 80


def create_win_loss_sheet(wb):
    """Create the Win/Loss by Competitor sheet."""
    ws = wb.create_sheet("Win-Loss by Competitor")

    set_column_widths(ws, [25, 15, 15, 20, 12, 15, 40, 20])

    headers = [
        "Deal Name",
        "Deal Size",
        "Close Date",
        "Competitor(s)",
        "Outcome",
        "Battle Card Used?",
        "What Worked / What Didn't",
        "Rep Name",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Sample data
    sample_data = [
        (
            "[Account Name]",
            "$50,000",
            "2025-01-15",
            "[Competitor A]",
            "Won",
            "Yes",
            "Win theme #2 resonated strongly. Landmine about uptime worked.",
            "[Rep Name]",
        ),
        (
            "[Account Name]",
            "$25,000",
            "2025-01-10",
            "[Competitor B]",
            "Lost",
            "No",
            "Lost on price. Need better TCO talking points.",
            "[Rep Name]",
        ),
        (
            "[Account Name]",
            "$100,000",
            "2025-01-08",
            "[Competitor A], [Competitor C]",
            "Won",
            "Yes",
            "Customer quote from case study was decisive.",
            "[Rep Name]",
        ),
    ]

    for row_num, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 45

    # Summary section
    row = 7
    ws.cell(row=row, column=1, value="SUMMARY METRICS").font = BOLD_FONT
    row += 1

    metrics = [
        ("Total Competitive Deals:", "[X]"),
        ("Win Rate (Overall):", "[X%]"),
        ("Win Rate vs [Competitor A]:", "[X%]"),
        ("Win Rate vs [Competitor B]:", "[X%]"),
        ("Deals Using Battle Card:", "[X]"),
        ("Win Rate with Battle Card:", "[X%]"),
        ("Win Rate without Battle Card:", "[X%]"),
    ]

    for metric, value in metrics:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        row += 1


def main():
    """Generate the battle card template Excel file."""
    wb = Workbook()

    create_battle_card_sheet(wb)
    create_competitor_tracker_sheet(wb)
    create_objection_library_sheet(wb)
    create_win_loss_sheet(wb)

    # Save to public/templates
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(
        script_dir, "..", "public", "templates", "battle-card-template.xlsx"
    )
    output_path = os.path.normpath(output_path)

    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
