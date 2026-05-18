#!/usr/bin/env python3
"""
Script to generate the Product Launch Checklist & Timeline Template Excel file.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)

# Border style
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def style_header_row(ws, row, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER


def style_section_header(ws, row, col, text):
    """Create a section header cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = SECTION_FILL
    cell.font = BOLD_FONT
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER
    return cell


def create_launch_timeline_sheet(wb):
    """Create the main Launch Timeline sheet."""
    ws = wb.active
    ws.title = "Launch Timeline"

    set_column_widths(ws, [15, 30, 50, 20, 20, 20])

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="PRODUCT LAUNCH CHECKLIST & TIMELINE")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    # Launch info
    ws.cell(row=2, column=1, value="Product/Feature:").font = BOLD_FONT
    ws.cell(row=2, column=2, value="[Product Name]")
    ws.cell(row=2, column=3, value="Launch Date:").font = BOLD_FONT
    ws.cell(row=2, column=4, value="[Date]")
    ws.cell(row=2, column=5, value="Owner:").font = BOLD_FONT
    ws.cell(row=2, column=6, value="[Name]")

    row = 4

    # Pre-Launch Section (T-30 days)
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "PRE-LAUNCH (T-30 to T-7 days)")
    row += 1

    headers = ["Task", "Description", "Owner", "Due Date", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = LIGHT_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    row += 1

    pre_launch_tasks = [
        (
            "Messaging & Positioning",
            "Finalize core messaging, value props, and positioning statement",
            "[Product Marketing]",
            "[T-25]",
            "Not Started",
            "",
        ),
        (
            "Target Personas",
            "Confirm target buyer personas and use cases",
            "[Product Marketing]",
            "[T-25]",
            "Not Started",
            "",
        ),
        (
            "Competitive Analysis",
            "Review competitive landscape and differentiation",
            "[Product Marketing]",
            "[T-25]",
            "Not Started",
            "",
        ),
        (
            "Blog Post Draft",
            "Write launch announcement blog post",
            "[Content Marketing]",
            "[T-20]",
            "Not Started",
            "",
        ),
        (
            "Email Campaigns",
            "Design email sequences for customers and prospects",
            "[Marketing]",
            "[T-20]",
            "Not Started",
            "",
        ),
        (
            "Sales Deck",
            "Create sales enablement deck with talking points",
            "[Product Marketing]",
            "[T-18]",
            "Not Started",
            "",
        ),
        (
            "Product Demo Video",
            "Record demo video or walkthrough",
            "[Product Marketing]",
            "[T-18]",
            "Not Started",
            "",
        ),
        (
            "Website Updates",
            "Update product pages, features, and pricing",
            "[Marketing]",
            "[T-15]",
            "Not Started",
            "",
        ),
        (
            "Sales Training",
            "Schedule and prepare sales team training session",
            "[Sales Enablement]",
            "[T-15]",
            "Not Started",
            "",
        ),
        (
            "Support Documentation",
            "Create help docs, FAQs, and support materials",
            "[Product/Support]",
            "[T-12]",
            "Not Started",
            "",
        ),
        (
            "Press Release",
            "Draft press release (if applicable)",
            "[PR/Marketing]",
            "[T-10]",
            "Not Started",
            "",
        ),
        (
            "Analyst Briefings",
            "Schedule briefings with industry analysts",
            "[Product Marketing]",
            "[T-10]",
            "Not Started",
            "",
        ),
    ]

    for task, desc, owner, due, status, notes in pre_launch_tasks:
        ws.cell(row=row, column=1, value=task).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=4, value=due).border = THIN_BORDER
        ws.cell(row=row, column=5, value=status).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Launch Week Section
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "LAUNCH WEEK (T-7 to Launch Day)")
    row += 1

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = LIGHT_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    row += 1

    launch_week_tasks = [
        (
            "Final Asset Review",
            "Review all assets for accuracy and brand consistency",
            "[Product Marketing]",
            "[T-7]",
            "Not Started",
            "",
        ),
        (
            "Sales Team Briefing",
            "Conduct sales team training and Q&A session",
            "[Sales Enablement]",
            "[T-5]",
            "Not Started",
            "",
        ),
        (
            "Customer Success Briefing",
            "Train CS team on new features and common questions",
            "[Customer Success]",
            "[T-5]",
            "Not Started",
            "",
        ),
        (
            "Support Team Briefing",
            "Train support team on troubleshooting and FAQs",
            "[Support]",
            "[T-5]",
            "Not Started",
            "",
        ),
        (
            "Blog Post Publish",
            "Publish launch announcement blog post",
            "[Content Marketing]",
            "[T-3]",
            "Not Started",
            "",
        ),
        (
            "Email Campaign Launch",
            "Send launch announcement emails to segments",
            "[Marketing]",
            "[T-2]",
            "Not Started",
            "",
        ),
        (
            "Social Media Posts",
            "Schedule and publish social media announcements",
            "[Marketing]",
            "[T-1]",
            "Not Started",
            "",
        ),
        (
            "Press Release Distribution",
            "Distribute press release to media (if applicable)",
            "[PR]",
            "[Launch Day]",
            "Not Started",
            "",
        ),
        (
            "Product Release",
            "Enable feature/product for all customers",
            "[Product/Engineering]",
            "[Launch Day]",
            "Not Started",
            "",
        ),
        (
            "Internal Announcement",
            "Announce launch to internal team",
            "[Product Marketing]",
            "[Launch Day]",
            "Not Started",
            "",
        ),
    ]

    for task, desc, owner, due, status, notes in launch_week_tasks:
        ws.cell(row=row, column=1, value=task).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=4, value=due).border = THIN_BORDER
        ws.cell(row=row, column=5, value=status).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Post-Launch Section
    ws.merge_cells(f"A{row}:F{row}")
    style_section_header(ws, row, 1, "POST-LAUNCH (Launch Day +1 to +30)")
    row += 1

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = LIGHT_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    row += 1

    post_launch_tasks = [
        (
            "Monitor Metrics",
            "Track adoption, engagement, and usage metrics",
            "[Product Marketing]",
            "[Ongoing]",
            "Not Started",
            "",
        ),
        (
            "Collect Feedback",
            "Gather feedback from early adopters and customers",
            "[Product/CS]",
            "[Ongoing]",
            "Not Started",
            "",
        ),
        (
            "Sales Follow-up",
            "Check in with sales team on deal conversations",
            "[Sales Enablement]",
            "[+3 days]",
            "Not Started",
            "",
        ),
        (
            "Support Ticket Review",
            "Review support tickets for common issues",
            "[Support]",
            "[+5 days]",
            "Not Started",
            "",
        ),
        (
            "Case Study Outreach",
            "Identify customers for case studies",
            "[Customer Marketing]",
            "[+7 days]",
            "Not Started",
            "",
        ),
        (
            "Iterate Messaging",
            "Refine messaging based on market feedback",
            "[Product Marketing]",
            "[+14 days]",
            "Not Started",
            "",
        ),
        (
            "Launch Retrospective",
            "Conduct post-launch retrospective meeting",
            "[Product Marketing]",
            "[+21 days]",
            "Not Started",
            "",
        ),
        (
            "Quarterly Review",
            "Review launch success metrics and learnings",
            "[Product Marketing]",
            "[+30 days]",
            "Not Started",
            "",
        ),
    ]

    for task, desc, owner, due, status, notes in post_launch_tasks:
        ws.cell(row=row, column=1, value=task).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=4, value=due).border = THIN_BORDER
        ws.cell(row=row, column=5, value=status).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1


def create_raci_matrix_sheet(wb):
    """Create the RACI Matrix sheet."""
    ws = wb.create_sheet("RACI Matrix")

    set_column_widths(ws, [30, 15, 15, 15, 15, 15, 15, 15, 40])

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value="RACI MATRIX - Cross-Functional Owners")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    # Legend
    ws.cell(row=row, column=1, value="R = Responsible").font = BOLD_FONT
    ws.cell(row=row, column=2, value="A = Accountable")
    ws.cell(row=row, column=3, value="C = Consulted")
    ws.cell(row=row, column=4, value="I = Informed")
    row += 2

    headers = [
        "Task/Activity",
        "Product Marketing",
        "Product",
        "Engineering",
        "Sales",
        "Marketing",
        "Support/CS",
        "PR",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    raci_tasks = [
        (
            "Messaging & Positioning",
            "A",
            "C",
            "",
            "C",
            "C",
            "",
            "",
            "Product Marketing owns, others provide input",
        ),
        (
            "Blog Post",
            "R",
            "C",
            "",
            "",
            "A",
            "",
            "C",
            "Marketing approves, Product Marketing writes",
        ),
        (
            "Sales Deck",
            "A",
            "C",
            "",
            "C",
            "",
            "",
            "",
            "Sales provides feedback",
        ),
        (
            "Product Demo Video",
            "R",
            "A",
            "C",
            "C",
            "",
            "",
            "",
            "Product owns, Product Marketing creates",
        ),
        (
            "Website Updates",
            "C",
            "C",
            "",
            "",
            "A",
            "",
            "",
            "Marketing owns execution",
        ),
        (
            "Sales Training",
            "A",
            "C",
            "",
            "C",
            "",
            "",
            "",
            "Sales Enablement supports",
        ),
        (
            "Support Documentation",
            "C",
            "A",
            "C",
            "",
            "",
            "C",
            "",
            "Product/Support own",
        ),
        (
            "Press Release",
            "C",
            "C",
            "",
            "",
            "C",
            "",
            "A",
            "PR owns, others review",
        ),
        (
            "Analyst Briefings",
            "A",
            "C",
            "",
            "",
            "",
            "",
            "",
            "Product Marketing leads",
        ),
        (
            "Customer Communications",
            "A",
            "",
            "",
            "",
            "R",
            "C",
            "",
            "Marketing executes, Product Marketing approves",
        ),
        (
            "Metrics Tracking",
            "A",
            "C",
            "C",
            "",
            "",
            "",
            "",
            "Product Marketing owns analysis",
        ),
    ]

    for task_data in raci_tasks:
        task = task_data[0]
        roles = task_data[1:8]
        notes = task_data[8]

        ws.cell(row=row, column=1, value=task).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

        for col, role in enumerate(roles, 2):
            cell = ws.cell(row=row, column=col, value=role)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if role == "A":
                cell.fill = PatternFill(
                    start_color="FFE699", end_color="FFE699", fill_type="solid"
                )
            elif role == "R":
                cell.fill = PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                )

        ws.cell(row=row, column=9, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1


def create_asset_checklist_sheet(wb):
    """Create the Asset Checklist sheet."""
    ws = wb.create_sheet("Asset Checklist")

    set_column_widths(ws, [30, 20, 20, 20, 15, 40])

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="ASSET CHECKLIST")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    headers = [
        "Asset",
        "Owner",
        "Due Date",
        "Status",
        "Published",
        "Notes/Link",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    assets = [
        (
            "Blog Post - Launch Announcement",
            "[Content Marketing]",
            "[T-3]",
            "Not Started",
            "",
            "",
        ),
        (
            "Blog Post - How It Works",
            "[Content Marketing]",
            "[T-7]",
            "Not Started",
            "",
            "",
        ),
        (
            "Blog Post - Use Cases",
            "[Content Marketing]",
            "[T-10]",
            "Not Started",
            "",
            "",
        ),
        (
            "Email - Customer Announcement",
            "[Marketing]",
            "[T-2]",
            "Not Started",
            "",
            "",
        ),
        (
            "Email - Prospect Nurture",
            "[Marketing]",
            "[T-2]",
            "Not Started",
            "",
            "",
        ),
        (
            "Sales Deck - Internal",
            "[Product Marketing]",
            "[T-18]",
            "Not Started",
            "",
            "",
        ),
        (
            "Sales One-Pager",
            "[Product Marketing]",
            "[T-18]",
            "Not Started",
            "",
            "",
        ),
        (
            "Product Demo Video",
            "[Product Marketing]",
            "[T-18]",
            "Not Started",
            "",
            "",
        ),
        (
            "Website - Product Page",
            "[Marketing]",
            "[T-15]",
            "Not Started",
            "",
            "",
        ),
        (
            "Website - Feature Page",
            "[Marketing]",
            "[T-15]",
            "Not Started",
            "",
            "",
        ),
        (
            "Website - Pricing Page",
            "[Marketing]",
            "[T-15]",
            "Not Started",
            "",
            "",
        ),
        (
            "Social Media - LinkedIn Post",
            "[Marketing]",
            "[T-1]",
            "Not Started",
            "",
            "",
        ),
        (
            "Social Media - Twitter/X Post",
            "[Marketing]",
            "[T-1]",
            "Not Started",
            "",
            "",
        ),
        (
            "Press Release",
            "[PR]",
            "[Launch Day]",
            "Not Started",
            "",
            "",
        ),
        (
            "Support Documentation",
            "[Product/Support]",
            "[T-12]",
            "Not Started",
            "",
            "",
        ),
        (
            "FAQ Document",
            "[Product/Support]",
            "[T-12]",
            "Not Started",
            "",
            "",
        ),
        (
            "Release Notes",
            "[Product]",
            "[Launch Day]",
            "Not Started",
            "",
            "",
        ),
        (
            "Analyst Briefing Deck",
            "[Product Marketing]",
            "[T-10]",
            "Not Started",
            "",
            "",
        ),
    ]

    for asset, owner, due, status, published, notes in assets:
        ws.cell(row=row, column=1, value=asset).border = THIN_BORDER
        ws.cell(row=row, column=2, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=3, value=due).border = THIN_BORDER
        ws.cell(row=row, column=4, value=status).border = THIN_BORDER
        ws.cell(row=row, column=5, value=published).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1


def create_metrics_tracker_sheet(wb):
    """Create the Metrics Tracker sheet."""
    ws = wb.create_sheet("Metrics Tracker")

    set_column_widths(ws, [25, 20, 20, 20, 20, 30])

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="LAUNCH METRICS TRACKER")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    row = 3

    headers = [
        "Metric",
        "Baseline (Pre-Launch)",
        "Week 1",
        "Week 2",
        "Week 4",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30
    row += 1

    metrics = [
        (
            "Product Adoption Rate",
            "",
            "",
            "",
            "",
            "% of eligible customers using feature",
        ),
        (
            "Active Users",
            "",
            "",
            "",
            "",
            "Number of users actively using feature",
        ),
        (
            "Feature Usage Frequency",
            "",
            "",
            "",
            "",
            "Average uses per user per week",
        ),
        (
            "Blog Post Views",
            "",
            "",
            "",
            "",
            "Total views on launch blog post",
        ),
        (
            "Email Open Rate",
            "",
            "",
            "",
            "",
            "Open rate for launch announcement emails",
        ),
        (
            "Email Click Rate",
            "",
            "",
            "",
            "",
            "Click rate for launch announcement emails",
        ),
        (
            "Sales Conversations",
            "",
            "",
            "",
            "",
            "Number of deals mentioning new feature",
        ),
        (
            "Support Tickets",
            "",
            "",
            "",
            "",
            "Support tickets related to new feature",
        ),
        (
            "Customer Feedback Score",
            "",
            "",
            "",
            "",
            "NPS or satisfaction score for feature",
        ),
        (
            "Social Media Engagement",
            "",
            "",
            "",
            "",
            "Likes, shares, comments on launch posts",
        ),
        (
            "Press Mentions",
            "",
            "",
            "",
            "",
            "Number of press/analyst mentions",
        ),
        (
            "Website Traffic",
            "",
            "",
            "",
            "",
            "Traffic to product/feature pages",
        ),
    ]

    for metric, baseline, w1, w2, w4, notes in metrics:
        ws.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=2, value=baseline).border = THIN_BORDER
        ws.cell(row=row, column=3, value=w1).border = THIN_BORDER
        ws.cell(row=row, column=4, value=w2).border = THIN_BORDER
        ws.cell(row=row, column=5, value=w4).border = THIN_BORDER
        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1


def main():
    """Generate the product launch checklist template Excel file."""
    wb = Workbook()

    create_launch_timeline_sheet(wb)
    create_raci_matrix_sheet(wb)
    create_asset_checklist_sheet(wb)
    create_metrics_tracker_sheet(wb)

    # Save to public/templates
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(
        script_dir,
        "..",
        "public",
        "templates",
        "product-launch-checklist-template.xlsx",
    )
    output_path = os.path.normpath(output_path)

    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
