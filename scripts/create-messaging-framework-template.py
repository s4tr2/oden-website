#!/usr/bin/env python3
"""
Generate the Messaging Framework Template Excel file.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_column_widths(ws, widths):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def style_table_header(ws, row, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_section_title(ws, cell_ref, value, merge_ref=None):
    if merge_ref:
        ws.merge_cells(merge_ref)
    cell = ws[cell_ref]
    cell.value = value
    cell.fill = SECTION_FILL
    cell.font = BOLD_FONT
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


def style_data_block(ws, start_row, rows):
    current_row = start_row
    for label, value in rows:
        ws.cell(row=current_row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=value)
        ws.cell(row=current_row, column=1).fill = LIGHT_FILL
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = THIN_BORDER
            ws.cell(row=current_row, column=col).alignment = Alignment(
                vertical="top", wrap_text=True
            )
        current_row += 1
    return current_row


def create_core_messaging_sheet(wb):
    ws = wb.active
    ws.title = "Core Messaging"
    set_column_widths(ws, [26, 70])

    ws.merge_cells("A1:B1")
    ws["A1"] = "MESSAGING FRAMEWORK - CORE MESSAGING"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Product / Offering"
    ws["B3"] = "[Product name]"
    ws["A4"] = "Last Updated"
    ws["B4"] = "[Date]"
    for row in (3, 4):
        ws[f"A{row}"].font = BOLD_FONT
        ws[f"A{row}"].fill = LIGHT_FILL
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"B{row}"].border = THIN_BORDER

    row = 6
    style_section_title(ws, f"A{row}", "FOUNDATION", f"A{row}:B{row}")
    row += 1
    row = style_data_block(
        ws,
        row,
        [
            ("Category / Market", "[What category are you in?]"),
            ("Ideal Customer Profile", "[Who is this for?]"),
            ("Primary Problem", "[What problem is most urgent?]"),
            ("Main Alternatives", "[What are buyers comparing you against?]"),
        ],
    )

    row += 1
    style_section_title(ws, f"A{row}", "POSITIONING", f"A{row}:B{row}")
    row += 1
    row = style_data_block(
        ws,
        row,
        [
            ("Positioning Statement", "[For..., who..., our product is..., that...]"),
            ("One-Line Value Proposition", "[Single sentence version for website or deck]"),
            ("Brand Promise", "[What outcome do customers get?]"),
            ("Top Differentiators", "[3 short bullets or phrases]"),
        ],
    )

    row += 1
    style_section_title(ws, f"A{row}", "CUSTOMER PAINS & OUTCOMES", f"A{row}:B{row}")
    row += 1
    style_data_block(
        ws,
        row,
        [
            ("Pain Point 1", "[Urgent pain or friction]"),
            ("Pain Point 2", "[Repeated pain]"),
            ("Pain Point 3", "[Strategic pain]"),
            ("Desired Outcome", "[What success looks like for the buyer]"),
        ],
    )


def create_value_pillars_sheet(wb):
    ws = wb.create_sheet("Value Pillars")
    set_column_widths(ws, [20, 32, 28, 28, 28, 32])
    ws["A1"] = "VALUE PILLARS"
    ws["A1"].font = Font(bold=True, size=14)
    style_table_header(
        ws,
        3,
        [
            "Pillar",
            "What It Means",
            "Customer Problem Solved",
            "Why It Matters",
            "Proof Points",
            "Example Copy",
        ],
    )

    examples = [
        [
            "Pillar 1",
            "[Short description]",
            "[Problem this solves]",
            "[Business or personal impact]",
            "[Metric, quote, review, case study]",
            "[How you might say this in copy]",
        ],
        ["Pillar 2", "", "", "", "", ""],
        ["Pillar 3", "", "", "", "", ""],
        ["Pillar 4", "", "", "", "", ""],
    ]

    for row_index, values in enumerate(examples, start=4):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 42

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:F7"


def create_persona_sheet(wb):
    ws = wb.create_sheet("Persona Messaging")
    set_column_widths(ws, [22, 24, 28, 28, 28, 26, 34])
    ws["A1"] = "PERSONA MESSAGING"
    ws["A1"].font = Font(bold=True, size=14)
    style_table_header(
        ws,
        3,
        [
            "Persona / Role",
            "Primary Goal",
            "Pain Points",
            "What They Care About",
            "Message Angle",
            "Likely Objections",
            "Recommended Talk Track",
        ],
    )

    rows = [
        [
            "[Primary buyer]",
            "[What they are trying to achieve]",
            "[Top pains]",
            "[Buying criteria]",
            "[How to frame value]",
            "[Common concerns]",
            "[Suggested talk track]",
        ],
        ["[Secondary buyer]", "", "", "", "", "", ""],
        ["[Executive sponsor]", "", "", "", "", "", ""],
    ]

    for row_index, values in enumerate(rows, start=4):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 48

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:G6"


def create_proof_points_sheet(wb):
    ws = wb.create_sheet("Proof Points")
    set_column_widths(ws, [20, 26, 24, 24, 20, 28])
    ws["A1"] = "PROOF POINTS & EVIDENCE"
    ws["A1"].font = Font(bold=True, size=14)
    style_table_header(
        ws,
        3,
        [
            "Proof Type",
            "Claim Supported",
            "Source",
            "Suggested Usage",
            "Audience Fit",
            "Notes",
        ],
    )

    rows = [
        [
            "Customer Quote",
            "[Which claim does this support?]",
            "[Customer / case study / call]",
            "[Homepage / sales deck / launch email]",
            "[Persona or segment]",
            "[Approval or freshness notes]",
        ],
        ["Metric", "", "", "", "", ""],
        ["Case Study", "", "", "", "", ""],
        ["Review / Analyst Note", "", "", "", "", ""],
    ]

    for row_index, values in enumerate(rows, start=4):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 42

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:F7"


def create_objection_handling_sheet(wb):
    ws = wb.create_sheet("Objection Handling")
    set_column_widths(ws, [26, 28, 34, 26, 22])
    ws["A1"] = "OBJECTION HANDLING"
    ws["A1"].font = Font(bold=True, size=14)
    style_table_header(
        ws,
        3,
        [
            "Objection",
            "What Is Behind It",
            "Recommended Response",
            "Supporting Evidence",
            "Best-Fit Persona / Segment",
        ],
    )

    rows = [
        [
            '"We already have a solution for this."',
            "[Switching cost, inertia, low urgency]",
            "[Acknowledge current state and reframe around missing outcome]",
            "[Customer result or comparison point]",
            "[Persona]",
        ],
        [
            '"Your product seems expensive."',
            "[Budget pressure or unclear ROI]",
            "[Connect price to outcome, efficiency, or avoided cost]",
            "[ROI metric or case study]",
            "[Persona]",
        ],
        [
            '"This looks hard to implement."',
            "[Perceived complexity or team capacity concerns]",
            "[Show time-to-value and onboarding support]",
            "[Implementation proof point]",
            "[Persona]",
        ],
    ]

    for row_index, values in enumerate(rows, start=4):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 54

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:E6"


def main():
    wb = Workbook()
    create_core_messaging_sheet(wb)
    create_value_pillars_sheet(wb)
    create_persona_sheet(wb)
    create_proof_points_sheet(wb)
    create_objection_handling_sheet(wb)

    output_path = os.path.join(
        os.path.dirname(__file__), "..", "public", "templates", "messaging-framework-template.xlsx"
    )
    wb.save(output_path)
    print(f"Created {output_path}")


if __name__ == "__main__":
    main()
